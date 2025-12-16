import os
import sys
import re
import json
import time
import sqlite3
import tempfile
from datetime import datetime, timezone
from urllib.parse import urlencode
from dateutil.parser import parse as dt_parse
from dateutil.tz import tzutc

import requests
import pandas as pd
from bs4 import BeautifulSoup
from msal import PublicClientApplication
import streamlit as st

from typing import Tuple

# =========================
# Paths / packaging support
# =========================
def base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = base_dir()

# Load environment variables from Streamlit secrets (cloud) or .env (local)
try:
    # Try Streamlit secrets first (for cloud deployment)
    GEMINI_API_KEY = st.secrets["credentials"]["GEMINI_API_KEY"]
    TENANT_ID = st.secrets["credentials"]["TENANT_ID"]
    CLIENT_ID = st.secrets["credentials"]["CLIENT_ID"]
    CLIENT_SECRET = st.secrets["credentials"].get("CLIENT_SECRET", "")
    MAILBOX = st.secrets["credentials"]["MAILBOX"]
    START_DATE = st.secrets["credentials"].get("START_DATE", "2025-01-01T00:00:00Z")
    PN_MASTER_PATH = st.secrets["credentials"].get("PN_MASTER_PATH", os.path.join(BASE_DIR, "pn_master.xlsx"))
except (KeyError, FileNotFoundError, AttributeError):
    # Fall back to .env for local development
    try:
        from dotenv import load_dotenv
        load_dotenv(os.path.join(BASE_DIR, ".env"))
    except:
        pass
    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
    TENANT_ID = os.getenv("TENANT_ID")
    CLIENT_ID = os.getenv("CLIENT_ID")
    CLIENT_SECRET = os.getenv("CLIENT_SECRET", "")
    START_DATE = os.getenv("START_DATE", "2025-01-01T00:00:00Z")
    MAILBOX = os.getenv("MAILBOX", "me")
    PN_MASTER_PATH = os.getenv("PN_MASTER_PATH", os.path.join(BASE_DIR, "pn_master.xlsx"))

# Quiet down gRPC / absl noise BEFORE importing google-generativeai
os.environ.setdefault("GRPC_VERBOSITY", "ERROR")
os.environ.setdefault("GRPC_LOG_SEVERITY", "ERROR")
os.environ.setdefault("GLOG_minloglevel", "2")
os.environ.setdefault("ABSL_LOG_SEVERITY", "info")

import google.generativeai as genai  # noqa: E402

DB_PATH = os.path.join(BASE_DIR, "complaints.db")
EXCEL_PATH = os.path.join(BASE_DIR, "Complaint_Log.xlsx")

def _mask_key(k: str) -> str:
    if not k:
        return "(empty)"
    if len(k) <= 8:
        return "*" * (len(k) - 2) + k[-2:]
    return k[:4] + "..." + k[-4:]

def _gemini_preflight(genai_mod, api_key: str, model_name: str = "gemini-2.0-flash") -> bool:
    """Check if Gemini API key works"""
    try:
        genai_mod.configure(api_key=api_key)
        model = genai_mod.GenerativeModel(
            model_name=model_name,
            generation_config={"response_mime_type": "application/json", "temperature": 0},
        )
        resp = model.generate_content("Return {\"ping\":true}")
        _ = getattr(resp, "text", None)
        return True
    except Exception as e:
        st.error(f"Gemini API key invalid or expired: {e}")
        return False

from prompts import PROMPT_TEXT

# ============================
# Complaint keyword gate list
# ============================
KEYWORDS = [
    "NCMR","rejection","defect","missing parts","damage","damaged","SCAR","DMR",
    "RTV","non-conformance","nonconformance","corrective action","cracking","reject","deficiency",
    "breaking","wrong revision","credit note","supplier corrective action request",
    "RMA","return","replacement","rework"
]

STRONG_SIGNALS = {"ncmr", "scar", "dmr", "rma", "nonconformance", "non-conformance", "ncr", "car", "8d"}

SENDER_BLOCKLIST = {
    "eminder@culturewise.com",
    "no-reply@culturewise.com",
}
DOMAIN_BLOCKLIST = {"culturewise.com"}
SUBJECT_BLOCK_PHRASES = {
    "lesson of the week",
    "reminder:",
    "training",
    "newsletter",
    "guide to best practices",
    "best practices",
    "weekly update",
    "out of office",
    "automatic reply",
}

MISSING_PN = "No part number provided"

# =================
# Graph Auth setup - STREAMLIT VERSION
# =================
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPES = ["Mail.Read", "User.Read"]
app = PublicClientApplication(CLIENT_ID, authority=AUTHORITY)

# ====================
# FIXED AUTHENTICATION FUNCTION
# Replace the get_token() function in your main.py with this version
# ====================

def get_token():
    """
    Streamlit-compatible authentication using device code flow.
    Shows the device code in the Streamlit UI with proper button layout.
    Stores token in session state for reuse.
    """
    # Check if we already have a valid token in session state
    if "access_token" in st.session_state and "token_expires_at" in st.session_state:
        if time.time() < st.session_state.token_expires_at:
            return st.session_state.access_token
    
    # Try silent authentication with cached account
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
        if result and "access_token" in result:
            st.session_state.access_token = result["access_token"]
            st.session_state.token_expires_at = time.time() + result.get("expires_in", 3600)
            return result["access_token"]
    
    # Need to show device code flow UI
    if "device_flow" not in st.session_state:
        # Start device code flow
        flow = app.initiate_device_flow(scopes=SCOPES)
        if "user_code" not in flow:
            st.error(f"Device code flow error: {flow}")
            st.stop()
        
        st.session_state.device_flow = flow
        st.session_state.auth_started = time.time()
    
    flow = st.session_state.device_flow
    
    # Display authentication UI
    st.warning("🔐 **Microsoft Authentication Required**")
    
    st.info("""
    **To sync emails, you need to authenticate with Microsoft:**
    
    1. Click "Open Login Page" below
    2. Enter the code shown
    3. Sign in with your work account
    4. Come back here and click "Check Authentication"
    """)
    
    # Show the device code prominently
    st.code(flow['user_code'], language=None)
    
    # Create buttons in a row - FIXED LAYOUT
    col1, col2 = st.columns(2)
    
    with col1:
        # Open login page button with proper link
        login_url = "https://microsoft.com/devicelogin"
        st.link_button("🌐 Open Login Page", login_url, use_container_width=True)
    
    with col2:
        # Check authentication button
        if st.button("✅ Check Authentication", use_container_width=True, type="primary"):
            with st.spinner("Verifying authentication..."):
                try:
                    result = app.acquire_token_by_device_flow(flow)
                    
                    if "access_token" in result:
                        st.session_state.access_token = result["access_token"]
                        st.session_state.token_expires_at = time.time() + result.get("expires_in", 3600)
                        st.session_state.pop("device_flow", None)
                        st.session_state.pop("auth_started", None)
                        st.success("✅ Authentication successful!")
                        time.sleep(1)
                        st.rerun()
                    else:
                        error_desc = result.get('error_description', 'Unknown error')
                        if "pending" in error_desc.lower():
                            st.warning("⏳ Still waiting for you to complete sign-in. Try again in a moment.")
                        else:
                            st.error(f"Authentication failed: {error_desc}")
                            # Reset flow on actual failure
                            st.session_state.pop("device_flow", None)
                except Exception as e:
                    st.error(f"Authentication error: {str(e)}")
    
    # Show timeout warning
    if "auth_started" in st.session_state:
        elapsed = time.time() - st.session_state.auth_started
        remaining = max(0, 900 - elapsed)  # 15 minute timeout
        
        if remaining > 0:
            minutes = int(remaining // 60)
            seconds = int(remaining % 60)
            st.caption(f"⏱️ Code expires in {minutes}m {seconds}s")
        else:
            st.error("⏰ Code expired. Click below to get a new code.")
            if st.button("🔄 Get New Code", use_container_width=True):
                st.session_state.pop("device_flow", None)
                st.session_state.pop("auth_started", None)
                st.rerun()
    
    # Stop execution until authenticated
    st.stop()

SUBJECT_PREFIXES = ("re:", "fw:", "fwd:", "sv:", "答复:", "回复:", "aw:", "wg:", "r:")

def clean_subject(subject: str) -> str:
    if not subject:
        return ""
    s = subject.strip()
    while True:
        low = s.lower().lstrip()
        if any(low.startswith(p) for p in SUBJECT_PREFIXES):
            colon_ix = low.find(":")
            s = s[colon_ix + 1:].strip()
        else:
            break
    s = re.sub(r"^(re|fw|fwd)[\s\-\:]+", "", s, flags=re.I).strip()
    return s

QUOTED_MARKERS = [
    "-----Original Message-----",
    "\nFrom:", "\r\nFrom:", "\nSent:", "\r\nSent:",
    "\nOn ", "\r\nOn ",
    "________________________________",
    "Forwarded message", "Original Appointment",
]

def to_et_naive(dt_utc_str: str):
    """Convert Graph ISO UTC string -> Eastern Time (ET), return a *naive* datetime"""
    try:
        from zoneinfo import ZoneInfo
        dt = datetime.fromisoformat(dt_utc_str.replace("Z", "+00:00"))
        dt_et = dt.astimezone(ZoneInfo("America/New_York"))
        return dt_et.replace(tzinfo=None)
    except Exception:
        return None

def trim_to_latest_reply(text: str) -> str:
    if not text:
        return ""
    cut = len(text)
    for m in QUOTED_MARKERS:
        idx = text.find(m)
        if idx != -1:
            cut = min(cut, idx)
    sep = re.search(r"\n-{5,}\n", text)
    if sep:
        cut = min(cut, sep.start())
    return text[:cut].strip()

def strip_html(html_text: str) -> str:
    if not html_text:
        return ""
    soup = BeautifulSoup(html_text, "html.parser")
    text = soup.get_text(" ", strip=True)
    return re.sub(r"\s+", " ", text).strip()

def contains_keywords(text: str) -> bool:
    low = text.lower()
    return any(kw.lower() in low for kw in KEYWORDS)

def is_noise_email(subject: str, sender: str, body: str) -> bool:
    s = (subject or "").lower()
    b = (body or "").lower()
    sender = (sender or "").lower()

    if sender in SENDER_BLOCKLIST:
        return True
    dom = sender.split("@")[-1] if "@" in sender else ""
    if dom in DOMAIN_BLOCKLIST:
        return True

    if any(phrase in s for phrase in SUBJECT_BLOCK_PHRASES):
        return True

    text = f"{s} {b}"
    has_keyword = any(kw.lower() in text for kw in KEYWORDS)
    has_strong = any(sig in text for sig in STRONG_SIGNALS)
    if not has_keyword and not has_strong:
        return True

    return False

# [INCLUDE ALL YOUR DATETIME AND ORIGIN EXTRACTION FUNCTIONS]
_SENT_LABELS = ("Sent:", "Date:", "Enviado:", "Fecha:", "Verzonden:", "Gesendet:")
_FALLBACK_DTFMTS = [
    "%A, %B %d, %Y %I:%M %p",
    "%a, %b %d, %Y %I:%M %p",
    "%m/%d/%Y %I:%M %p",
    "%Y-%m-%d %H:%M",
]
_EMAIL_ANYWHERE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I)

def _parse_human_datetime_to_utc_iso(s: str) -> str:
    if not s:
        return ""
    s = s.strip()
    try:
        from dateutil import parser
        from zoneinfo import ZoneInfo
        dt = parser.parse(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("America/New_York"))
        return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
    except Exception:
        pass
    try:
        from zoneinfo import ZoneInfo
    except Exception:
        ZoneInfo = None
    for fmt in _FALLBACK_DTFMTS:
        try:
            dt = datetime.strptime(s, fmt)
            if ZoneInfo:
                dt = dt.replace(tzinfo=ZoneInfo("America/New_York"))
            else:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
        except Exception:
            continue
    return ""

def _iso_to_dt(s: str):
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None

def _min_iso(*values: str) -> str:
    best = None
    for v in values:
        if not v:
            continue
        dv = _iso_to_dt(v)
        if dv is None:
            continue
        if best is None or dv < best:
            best = dv
    return best.astimezone(timezone.utc).isoformat().replace("+00:00", "Z") if best else ""

def extract_origins_deep(full_body_plain: str):
    if not full_body_plain:
        return "", ""
    lines = full_body_plain.splitlines()
    candidates = []
    for i, line in enumerate(lines):
        if "From:" in line:
            m_from = re.search(r"\bFrom:\s*(.*)", line, flags=re.I)
            if not m_from:
                continue
            emails = _EMAIL_ANYWHERE.findall(line)
            email = emails[0].strip() if emails else ""
            if not email:
                continue
            for j in range(i + 1, min(i + 10, len(lines))):
                l2 = lines[j]
                if any(l2.strip().lower().startswith(lbl.lower()) for lbl in _SENT_LABELS):
                    sent_text = l2.split(":", 1)[1] if ":" in l2 else l2
                    iso = _parse_human_datetime_to_utc_iso(sent_text)
                    if iso:
                        candidates.append((email, iso))
                        break
    inline = re.findall(
        r"^\s*On\s+(.+?)\s+(?:wrote|escribió):.*?([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})",
        full_body_plain, flags=re.I | re.M
    )
    for date_part, email in inline:
        iso = _parse_human_datetime_to_utc_iso(date_part)
        if iso:
            candidates.append((email.strip(), iso))
    if not candidates:
        return "", ""
    email, iso = min(candidates, key=lambda t: _iso_to_dt(t[1]) or datetime.max)
    return email, iso

def extract_earliest_datetime_anywhere(text: str) -> str:
    if not text:
        return ""
    candidates = set()
    for m in re.finditer(r"\b\d{4}-\d{1,2}-\d{1,2}(?:[ T]\d{1,2}:\d{2}(?::\d{2})?(?:\s?[AP]M)?)?\b", text):
        candidates.add(m.group(0))
    for m in re.finditer(r"\b\d{1,2}/\d{1,2}/\d{2,4}(?:\s+\d{1,2}:\d{2}(?::\d{2})?\s?(?:AM|PM)?)?\b", text, flags=re.I):
        candidates.add(m.group(0))
    for m in re.finditer(r"\b([A-Z][a-z]+)\s+\d{1,2},\s+\d{4}(?:\s+\d{1,2}:\d{2}(?::\d{2})?\s?(?:AM|PM)?)?\b", text):
        candidates.add(m.group(0))
    for m in re.finditer(r"\bOn\s+(.+?)\s+(?:wrote|escribió):", text, flags=re.I):
        candidates.add(m.group(1))
    if not candidates:
        return ""
    iso_values = []
    for tok in candidates:
        iso = _parse_human_datetime_to_utc_iso(tok)
        if iso:
            iso_values.append(iso)
    if not iso_values:
        return ""
    return _min_iso(*iso_values)

def _to_utc_iso_from_sent(s: str) -> str:
    return _parse_human_datetime_to_utc_iso(s)

def extract_origin_from_history(full_body_plain: str):
    if not full_body_plain:
        return "", ""
    lines = full_body_plain.splitlines()
    origin_email, origin_sent_iso = "", ""
    for i, line in enumerate(lines):
        m_from = re.match(r"\s*From:\s.*?<([^>\s@]+@[^>]+)>", line, flags=re.I)
        if not m_from:
            continue
        candidate_email = m_from.group(1).strip()
        for j in range(i + 1, min(i + 9, len(lines))):
            m_sent = re.match(r"\s*Sent:\s*(.*)$", lines[j], flags=re.I)
            if m_sent:
                candidate_sent = m_sent.group(1).strip()
                candidate_iso = _to_utc_iso_from_sent(candidate_sent)
                if candidate_iso:
                    origin_email = candidate_email
                    origin_sent_iso = candidate_iso
                break
    return origin_email, origin_sent_iso

# [INCLUDE ALL YOUR CASE KEY AND PN FUNCTIONS]
CASE_ID_PATTERNS = [
    (r'\bNCMR[\s:-]*([0-9]{4}[-/][0-9]{3,6}|[0-9]{6,})', 'ncmr'),
    (r'\bSCAR[\s:-]*([0-9]{4}[-/][0-9]{3,6}|[0-9]{5,})', 'scar'),
    (r'\bDMR[\s:-]*([0-9]{4}[-/][0-9]{3,6}|[0-9]{5,})', 'dmr'),
    (r'\bNCR[\s:-]*([0-9]{3,})', 'ncr'),
    (r'\bCAR[\s:-]*([0-9]{3,})', 'car'),
    (r'\bPO\s*(?:#|No\.?|Number)?\s*[:\-]?\s*([0-9]{5,})', 'po'),
    (r'\bSO\s*(?:#|No\.?|Number)?\s*[:\-]?\s*([0-9]{5,})', 'so'),
]

def compute_first_seen_initiator(conversation_id, full_body_plain, fallback_sender, mailbox):
    pattern = re.compile(
        r"From:\s*(.+?)<\s*([\w\.-]+@[\w\.-]+)\s*>.*?Sent:\s*([A-Za-z0-9,:\s\-]+(?:AM|PM)?)",
        re.IGNORECASE
    )
    matches = pattern.findall(full_body_plain)
    earliest_dt = None
    initiator_email = None
    for _, email, sent_raw in matches:
        try:
            sent_dt = dt_parse(sent_raw.strip(), fuzzy=True)
            if sent_dt.tzinfo is None:
                sent_dt = sent_dt.replace(tzinfo=tzutc())
            else:
                sent_dt = sent_dt.astimezone(tzutc())
            if not earliest_dt or sent_dt < earliest_dt:
                earliest_dt = sent_dt
                initiator_email = email.strip()
        except Exception:
            continue
    iso = earliest_dt.isoformat() if earliest_dt else ""
    return iso, initiator_email or fallback_sender

def extract_external_id(text: str) -> str:
    t = text or ""
    for pat, tag in CASE_ID_PATTERNS:
        m = re.search(pat, t, flags=re.I)
        if m:
            raw = m.group(1)
            norm = raw.replace(" ", "").replace("/", "-")
            return f"{tag}-{norm.lower()}"
    return ""

def normalize_case_key(raw: str) -> str:
    if not raw:
        return ""
    s = re.sub(r'[^a-z0-9\-_\/]','', raw.lower())
    return s[:80]

def canonical_case_key(domain: str, pn_norm: str, subject: str, text_for_ids: str) -> str:
    dom = (domain or "").lower()
    ext = extract_external_id(f"{subject or ''} {text_for_ids or ''}")
    if pn_norm:
        key = f"{dom}-{pn_norm}-{ext}" if ext else f"{dom}-{pn_norm}"
    elif ext:
        key = f"{dom}-{ext}"
    else:
        subj = re.sub(r'[^a-z0-9]+', '-', (subject or '').lower()).strip('-')[:30]
        key = f"{dom}-{subj or 'no-subject'}"
    return normalize_case_key(key)

def to_et(dt_utc_str: str) -> str:
    try:
        from zoneinfo import ZoneInfo
        et = ZoneInfo("America/New_York")
    except Exception:
        et = None
    try:
        dt = datetime.fromisoformat(dt_utc_str.replace("Z", "+00:00"))
        if et:
            dt = dt.astimezone(et)
        else:
            dt = dt.astimezone(timezone.utc)
    except Exception:
        return dt_utc_str
    return dt.strftime("%-I:%M %p %m/%d/%Y") if os.name != "nt" else dt.strftime("%#I:%M %p %m/%d/%Y")

# [DATABASE FUNCTIONS]
def touch_conversation(conversation_id: str, received_utc: str):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("UPDATE complaints SET received_utc=? WHERE conversation_id=?", (received_utc, conversation_id))
    con.commit()
    con.close()

def get_by_conversation_id(conv_id: str):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
        SELECT conversation_id, received_utc, part_number
        FROM complaints
        WHERE conversation_id=?
    """, (conv_id,))
    row = cur.fetchone()
    con.close()
    return row

def get_by_case_key(case_key: str):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
        SELECT conversation_id, received_utc, part_number
        FROM complaints
        WHERE case_key=?
        ORDER BY received_utc DESC
        LIMIT 1
    """, (case_key,))
    row = cur.fetchone()
    con.close()
    return row

def ensure_columns():
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("PRAGMA table_info(complaints)")
    cols = {row[1] for row in cur.fetchall()}
    if "first_seen_utc" not in cols:
        cur.execute("ALTER TABLE complaints ADD COLUMN first_seen_utc TEXT")
    if "initiator_email" not in cols:
        cur.execute("ALTER TABLE complaints ADD COLUMN initiator_email TEXT")
    con.commit()
    con.close()

def update_row_for_conversation(target_conv_id: str, row: dict):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
        UPDATE complaints SET
            received_utc = :received_utc,
            from_email = :from_email,
            subject = :subject,
            jo_number = :jo_number,
            part_number = :part_number,
            category = :category,
            summary = :summary,
            case_key = :case_key,
            thread_url = :thread_url,
            first_seen_utc = CASE
                WHEN first_seen_utc IS NULL THEN :first_seen_utc
                WHEN :first_seen_utc IS NOT NULL AND :first_seen_utc < first_seen_utc THEN :first_seen_utc
                ELSE first_seen_utc
            END,
            initiator_email = COALESCE(initiator_email, :initiator_email)
        WHERE conversation_id = :target_conv_id
    """, {**row, "target_conv_id": target_conv_id})
    con.commit()
    con.close()

# [PART NUMBER EXTRACTION]
PN_PATTERNS = [
    r'\bP\s*/?\s*N\s*(?:No\.?|#)?\s*[:\-]?\s*([A-Za-z0-9\-_\.\/]{5,25})',
    r'\b(Part|Item|SKU)\s*(?:No\.?|Number|#)?\s*[:\-]?\s*([A-Za-z0-9\-_\.\/]{5,25})',
    r'\bPN#?\s*[:\-]?\s*([A-Za-z0-9\-_\.\/]{5,25})',
]

STOPWORDS = {
    "or","and","ok","re","fw","bs","hn","hi","thanks","regards",
    "am","pm","to","on","by","in","the","for","of","a","an","it","is"
}

PN_ALLOWED = re.compile(r'^[A-Za-z0-9\-_\.\/]{5,25}$')

def _has_digit(s: str) -> bool:
    return any(ch.isdigit() for ch in s)

def _has_letter(s: str) -> bool:
    return any(ch.isalpha() for ch in s)

def is_valid_pn_basic(token: str) -> bool:
    if not token or token.lower() in STOPWORDS:
        return False
    if not PN_ALLOWED.match(token):
        return False
    return _has_digit(token) and _has_letter(token)

def normalize_pn(s: str) -> str:
    if not s:
        return ""
    s = s.upper().strip()
    s = s.replace(" ", "")
    return re.sub(r"[^A-Z0-9\-\_\.\/]", "", s)

def _alnum(s: str) -> str:
    return re.sub(r'[^A-Z0-9]', '', (s or '').upper())

def extract_pn_candidates(subject: str, latest_reply: str):
    hay = "  ".join([(subject or ""), (latest_reply or "")])
    master_hit = None
    fallback = None
    hits = []
    for pat in PN_PATTERNS:
        for m in re.finditer(pat, hay, flags=re.I):
            g = m.group(2) if m.lastindex and m.lastindex >= 2 else m.group(1)
            token = (g or "").strip().strip('.,;:)]}')
            if token and is_valid_pn_basic(token):
                hits.append((token, normalize_pn(token)))
    for token, norm in hits:
        if PN_MASTER_SET and norm in PN_MASTER_SET:
            master_hit = master_hit or token
        else:
            fallback = fallback or token
    return master_hit, fallback, hay, _alnum(hay)

def load_master_pns(path: str) -> set:
    if not os.path.exists(path):
        print(f"[WARN] PN master file not found at {path}")
        return set()
    try:
        if path.lower().endswith((".xlsx", ".xls")):
            df = pd.read_excel(path, engine="openpyxl")
        elif path.lower().endswith(".csv"):
            df = pd.read_csv(path)
        else:
            df = pd.read_excel(path, engine="openpyxl")
    except Exception as e:
        print(f"[WARN] Failed to load PN master file: {e}")
        return set()
    cols = [c for c in df.columns if str(c).strip().lower() in {"partnumber", "pn", "part", "item", "sku"}]
    col = cols[0] if cols else df.columns[0]
    values = (
        df[col].astype(str).map(normalize_pn).dropna().map(str.strip)
        .replace("", pd.NA).dropna().unique().tolist()
    )
    pnset = set(values)
    print(f"[INFO] Loaded {len(pnset)} master PNs from {path}")
    return pnset

PN_MASTER_SET = load_master_pns(PN_MASTER_PATH)

# [DATABASE INIT AND UPSERT]
def init_db():
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS complaints (
            conversation_id TEXT PRIMARY KEY,
            received_utc TEXT,
            from_email TEXT,
            subject TEXT,
            jo_number TEXT,
            part_number TEXT,
            category TEXT,
            summary TEXT,
            case_key TEXT,
            thread_url TEXT,
            first_seen_utc TEXT,
            initiator_email TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS custom_columns (
            column_name TEXT PRIMARY KEY,
            column_type TEXT DEFAULT 'TEXT'
        )
    """)
    con.commit()
    con.close()
    ensure_columns()

def upsert_row(row: dict):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
        INSERT INTO complaints (
            conversation_id, received_utc, from_email, subject, jo_number, part_number,
            category, summary, case_key, thread_url, first_seen_utc, initiator_email
        )
        VALUES (
            :conversation_id, :received_utc, :from_email, :subject, :jo_number, :part_number,
            :category, :summary, :case_key, :thread_url, :first_seen_utc, :initiator_email
        )
        ON CONFLICT(conversation_id) DO UPDATE SET
            received_utc = excluded.received_utc,
            from_email   = excluded.from_email,
            subject      = excluded.subject,
            jo_number    = excluded.jo_number,
            part_number  = excluded.part_number,
            category     = excluded.category,
            summary      = excluded.summary,
            case_key     = excluded.case_key,
            thread_url   = excluded.thread_url,
            first_seen_utc = CASE
                WHEN complaints.first_seen_utc IS NULL THEN excluded.first_seen_utc
                WHEN excluded.first_seen_utc IS NOT NULL AND excluded.first_seen_utc < complaints.first_seen_utc THEN excluded.first_seen_utc
                ELSE complaints.first_seen_utc
            END,
            initiator_email = COALESCE(complaints.initiator_email, excluded.initiator_email)
    """, row)
    con.commit()
    con.close()

def fetch_all_rows():
    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM complaints", con)
    con.close()
    return df

def _safe_write_excel(write_fn, target_path: str, retries: int = 3, sleep_s: float = 1.2):
    target_dir = os.path.dirname(os.path.abspath(target_path))
    base, ext = os.path.splitext(os.path.basename(target_path))
    tmp_fd, tmp_path = tempfile.mkstemp(prefix=base + "_tmp_", suffix=ext, dir=target_dir)
    os.close(tmp_fd)
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            write_fn(tmp_path)
            os.replace(tmp_path, target_path)
            print(f"[OK] Excel written: {target_path}")
            return target_path
        except PermissionError as e:
            last_err = e
            if attempt < retries:
                print(f"[WARN] Excel locked (attempt {attempt}/{retries}). Retrying...")
                time.sleep(sleep_s)
            else:
                print(f"[ERROR] Could not overwrite locked file: {target_path}")
        except Exception as e:
            last_err = e
            try: os.remove(tmp_path)
            except: pass
            raise
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    fallback_path = os.path.join(target_dir, f"{base}_{ts}{ext}")
    write_fn(fallback_path)
    print(f"[OK] Excel written (fallback): {fallback_path}")
    return fallback_path

def export_to_excel():
    df = fetch_all_rows()
    if df.empty:
        print("[INFO] No rows to export.")
        return
    if "case_key" in df.columns and "received_utc" in df.columns:
        df = df.sort_values("received_utc").drop_duplicates(subset=["case_key"], keep="last")
    def _to_et_naive_wrapper(x):
        return to_et_naive(x) if pd.notna(x) and str(x).strip() else None
    df["__first_et"] = df["first_seen_utc"].apply(_to_et_naive_wrapper)
    df["Date (ET)"] = df["__first_et"]
    df = df.sort_values("__first_et", ascending=False, na_position="last")
    colmap = {
        "initiator_email": "Initiated By",
        "part_number": "P/N",
        "summary": "Summary",
        "category": "Category",
        "subject": "Subject",
        "thread_url": "Link",
    }
    base_cols = ["Date (ET)"] + list(colmap.keys())
    base_cols = [c for c in base_cols if c in df.columns]
    df_out = df[base_cols].rename(columns=colmap)
    if "Category_Final" not in df_out.columns:
        if "Category" in df_out.columns:
            df_out.insert(df_out.columns.get_loc("Category") + 1, "Category_Final", "")
        else:
            df_out["Category_Final"] = ""
    if "Notes" not in df_out.columns:
        if "Subject" in df_out.columns:
            df_out.insert(df_out.columns.get_loc("Subject") + 1, "Notes", "")
        else:
            df_out["Notes"] = ""
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("SELECT column_name FROM custom_columns")
    custom_cols = [row[0] for row in cur.fetchall()]
    con.close()
    for col in custom_cols:
        if col in df.columns and col not in df_out.columns:
            df_out[col] = df[col]
    if "Link" in df_out.columns:
        df_out["_url"] = df_out["Link"].copy()
        df_out["Link"] = df_out["_url"].apply(
            lambda u: f'=HYPERLINK("{str(u).strip()}", "Open")' if pd.notna(u) and str(u).strip() else ""
        )
    desired = [
        "Date (ET)",
        "Initiated By",
        "P/N",
        "Category", "Category_Final",
        "Summary",
        "Subject",
        "Notes",
        "Link",
    ] + custom_cols
    existing = [c for c in desired if c in df_out.columns]
    if "_url" in df_out.columns:
        df_out = df_out.drop(columns=["_url"])
    df_out = df_out[existing]
    def _write(path):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            sheet = "Complaints"
            df_out.to_excel(writer, sheet_name=sheet, index=False)
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.styles import Alignment, Font
            wb = writer.book
            ws = wb[sheet]
            ws.freeze_panes = "A2"
            last_col = get_column_letter(ws.max_column)
            last_row = ws.max_row
            tbl = Table(displayName="ComplaintTable", ref=f"A1:{last_col}{last_row}")
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(tbl)
            if "Summary" in df_out.columns:
                cidx = df_out.columns.get_loc("Summary") + 1
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=cidx).alignment = Alignment(wrap_text=True, vertical="top")
            if "Date (ET)" in df_out.columns:
                didx = df_out.columns.get_loc("Date (ET)") + 1
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=didx).number_format = "mm/dd/yyyy"
            for c in range(1, ws.max_column + 1):
                ws.cell(row=1, column=c).font = Font(bold=True)
            target_widths = {
                "Date (ET)": 12,
                "Initiated By": 30,
                "P/N": 26,
                "Summary": 64,
                "Category": 18, "Category_Final": 18,
                "Subject": 44,
                "Notes": 30,
                "Link": 10,
            }
            for idx, name in enumerate(df_out.columns, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = target_widths.get(name, 24)
    _safe_write_excel(_write, EXCEL_PATH)

# [MICROSOFT GRAPH]
GRAPH_BASE = "https://graph.microsoft.com/v1.0"

def graph_headers(token: str):
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

def fetch_messages_since(start_iso: str, mailbox: str = MAILBOX, page_size: int = 50):
    token = get_token()
    if mailbox and mailbox != "me":
        base = f"{GRAPH_BASE}/users/{mailbox}/messages"
    else:
        base = f"{GRAPH_BASE}/me/messages"
    params = {
        "$top": page_size,
        "$orderby": "receivedDateTime asc",
        "$filter": f"receivedDateTime ge {start_iso}",
        "$select": "id,conversationId,receivedDateTime,subject,from,body,webLink,bodyPreview,internetMessageId"
    }
    url = f"{base}?{urlencode(params)}"
    while True:
        resp = requests.get(url, headers=graph_headers(token))
        if resp.status_code == 401:
            token = get_token()
            resp = requests.get(url, headers=graph_headers(token))
        if resp.status_code != 200:
            raise RuntimeError(f"Graph error {resp.status_code}: {resp.text}")
        data = resp.json()
        for item in data.get("value", []):
            yield item
        next_link = data.get("@odata.nextLink")
        if not next_link:
            break
        url = next_link

def fetch_earliest_in_conversation(conversation_id: str, mailbox: str = MAILBOX):
    token = get_token()
    if mailbox and mailbox != "me":
        base = f"{GRAPH_BASE}/users/{mailbox}/messages"
    else:
        base = f"{GRAPH_BASE}/me/messages"
    params = {
        "$top": 1,
        "$orderby": "receivedDateTime asc",
        "$filter": f"conversationId eq '{conversation_id}'",
        "$select": "id,receivedDateTime,from"
    }
    url = f"{base}?{urlencode(params)}"
    resp = requests.get(url, headers=graph_headers(token))
    if resp.status_code == 401:
        token = get_token()
        resp = requests.get(url, headers=graph_headers(token))
    if resp.status_code != 200:
        raise RuntimeError(f"Graph (earliest) error {resp.status_code}: {resp.text}")
    vals = resp.json().get("value", [])
    return vals[0] if vals else None

# [GEMINI]
def gemini_client():
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY not set")
    print("[CONFIG] Loaded GEMINI_API_KEY:", _mask_key(GEMINI_API_KEY))
    if not _gemini_preflight(genai, GEMINI_API_KEY, "gemini-2.0-flash"):
        raise SystemExit(1)
    genai.configure(api_key=GEMINI_API_KEY)
    return genai.GenerativeModel(
        model_name="gemini-2.0-flash",
        generation_config={
            "response_mime_type": "application/json",
            "temperature": 0,
        }
    )

def gemini_extract(model, subject_clean: str, from_email: str, latest_reply: str,
                   timeout_s: int = 30, retries: int = 3, backoff: float = 1.8) -> dict:
    prompt = PROMPT_TEXT.format(
        subject_clean=subject_clean,
        from_email=from_email,
        body_text=latest_reply
    )
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            resp = model.generate_content(prompt, request_options={"timeout": timeout_s})
            text = getattr(resp, "text", None)
            if not text:
                try:
                    text = resp.candidates[0].content.parts[0].text
                except Exception:
                    text = "{}"
            try:
                data = json.loads(text)
            except Exception:
                m = re.search(r"\{.*\}", text, flags=re.S)
                data = json.loads(m.group(0)) if m else {}
            if isinstance(data, list):
                if len(data)==1 and isinstance(data[0], dict):
                    data = data[0]
                else:
                    try:
                        data = dict(data)
                    except Exception:
                        data = {}
            if not isinstance(data, dict):
                data = {}
            return {
                "is_complaint": bool(data.get("is_complaint", False)),
                "summary": data.get("summary", ""),
                "category_suggested": data.get("category_suggested", "Other"),
                "case_key": data.get("case_key", ""),
                "part_number": data.get("part_number", ""),
            }
        except Exception as e:
            last_exc = e
            if attempt < retries:
                sleep_for = backoff ** (attempt - 1)
                print(f"[WARN] Gemini call failed (attempt {attempt}/{retries}): {e}")
                time.sleep(sleep_for)
            else:
                print(f"[ERROR] Gemini failed after {retries} attempts: {e}")
                break
    return {"is_complaint": False, "summary": "", "category_suggested": "Other", "case_key": "", "part_number": ""}

def tighten_summary(s: str, max_words=45):
    words = (s or "").split()
    return " ".join(words[:max_words])

CATEGORIES = [
    "Product","Shipping","Documentation/Revision","Invoicing/RTV",
    "Supplier/SCAR","Damage/Transit","Missing Parts","Other"
]

# [MAIN PROCESS FUNCTION]
def process(override_start_date=None):
    init_db()
    checked = 0
    new_threads = 0
    updated_threads = 0
    filtered_out = 0
    unchanged = 0
    updates_log = []
    latest_msg_by_conv = {}
    first_msg_by_conv = {}
    
    print("[INFO] Fetching messages from Graph...")
    start_iso = override_start_date or START_DATE
    for msg in fetch_messages_since(start_iso, MAILBOX):
        checked += 1
        conv_id = msg.get("conversationId")
        rdt = msg.get("receivedDateTime")
        if not conv_id or not rdt:
            continue
        if conv_id not in first_msg_by_conv:
            first_msg_by_conv[conv_id] = msg
        prev = latest_msg_by_conv.get(conv_id)
        if (not prev) or (rdt > prev.get("receivedDateTime", "")):
            latest_msg_by_conv[conv_id] = msg
    
    print(f"[INFO] Checked {checked} messages. Found {len(latest_msg_by_conv)} unique threads.")
    model = gemini_client()
    
    for conv_id, msg in latest_msg_by_conv.items():
        subject_raw = msg.get("subject") or ""
        subject_clean = clean_subject(subject_raw)
        sender_email = ((msg.get("from") or {}).get("emailAddress") or {}).get("address", "")
        rdt = msg.get("receivedDateTime") or ""
        body = msg.get("body") or {}
        body_content = body.get("content") or ""
        content_type = body.get("contentType") or "text"
        body_plain = strip_html(body_content) if content_type.lower() == "html" else body_content
        latest_reply = trim_to_latest_reply(body_plain)
        tail_text = body_plain[len(latest_reply):].strip() if len(body_plain) > len(latest_reply) else ""
        
        existing_by_conv = get_by_conversation_id(conv_id)
        if existing_by_conv:
            existing_rdt = existing_by_conv[1] or ""
            if existing_rdt and existing_rdt >= rdt:
                unchanged += 1
                continue
        
        if is_noise_email(subject_clean, sender_email, latest_reply):
            if existing_by_conv:
                touch_conversation(conv_id, rdt)
            filtered_out += 1
            continue
        
        llm_out = gemini_extract(model, subject_clean, sender_email, body_plain)
        if not llm_out.get("is_complaint", False):
            if existing_by_conv:
                touch_conversation(conv_id, rdt)
            filtered_out += 1
            continue
        
        summary = tighten_summary(llm_out.get("summary", ""))
        cat = llm_out.get("category_suggested", "Other")
        category_suggested = cat if cat in CATEGORIES else "Other"
        
        pn_master, pn_fallback, _hay_raw1, hay_alnum1 = extract_pn_candidates(subject_clean, latest_reply)
        pn_master2 = pn_fallback2 = None
        hay_alnum2 = ""
        if not pn_master and not pn_fallback and tail_text:
            pn_master2, pn_fallback2, _hay_raw2, hay_alnum2 = extract_pn_candidates("", tail_text)
        
        pn_final = pn_master or pn_master2 or pn_fallback or pn_fallback2
        
        if not pn_final:
            llm_pn = llm_out.get("part_number")
            if llm_pn and is_valid_pn_basic(llm_pn):
                llm_norm = normalize_pn(llm_pn)
                alnum_llm = _alnum(llm_pn)
                if (PN_MASTER_SET and llm_norm in PN_MASTER_SET) or (alnum_llm in (hay_alnum1 or "") or alnum_llm in (hay_alnum2 or "")):
                    pn_final = llm_pn
        
        if not pn_final:
            pn_final = MISSING_PN
        
        anywhere_iso = extract_earliest_datetime_anywhere(body_plain)
        helper_iso, helper_sender = compute_first_seen_initiator(
            conversation_id=conv_id,
            full_body_plain=body_plain,
            fallback_sender=sender_email,
            mailbox=MAILBOX,
        )
        
        first_seen_utc = _min_iso(anywhere_iso, helper_iso, rdt)
        initiator_email = helper_sender or sender_email
        first_seen_utc = first_seen_utc or rdt
        
        domain = (sender_email or "").split("@")[-1]
        pn_norm = normalize_pn(pn_final)
        case_key = canonical_case_key(
            domain=domain,
            pn_norm=pn_norm,
            subject=subject_clean,
            text_for_ids=f"{latest_reply}\n{summary}"
        )
        
        thread_url = msg.get("webLink") or ""
        
        if not existing_by_conv:
            try:
                earliest = fetch_earliest_in_conversation(conv_id)
                if earliest:
                    graph_api_earliest_iso = earliest.get("receivedDateTime", "")
                    first_seen_utc = _min_iso(first_seen_utc, graph_api_earliest_iso)
                    if not initiator_email:
                        initiator_email = (
                            ((earliest.get("from") or {}).get("emailAddress") or {}).get("address", "")
                            or initiator_email
                        )
            except Exception:
                pass
        
        row = {
            "conversation_id": conv_id,
            "received_utc": rdt,
            "from_email": sender_email,
            "subject": subject_clean,
            "jo_number": None,
            "part_number": pn_final,
            "category": category_suggested,
            "summary": summary,
            "case_key": case_key,
            "thread_url": thread_url,
            "first_seen_utc": first_seen_utc,
            "initiator_email": initiator_email,
        }
        
        existing_by_case = get_by_case_key(case_key)
        
        if existing_by_conv:
            upsert_row(row)
            updated_threads += 1
            prev_pn = (existing_by_conv[2] or "").strip()
            if prev_pn == MISSING_PN and pn_final != MISSING_PN:
                updates_log.append(f"Updated thread (PN captured): {subject_clean} (PN: {pn_final})")
            else:
                updates_log.append(f"Updated thread: {subject_clean} (PN: {pn_final})")
        elif existing_by_case:
            target_conv, target_rdt, target_pn = existing_by_case
            if (target_rdt or "") >= rdt:
                unchanged += 1
                continue
            update_row_for_conversation(target_conv, row)
            updated_threads += 1
            prev_pn = (target_pn or "").strip()
            if prev_pn == MISSING_PN and pn_final != MISSING_PN:
                updates_log.append(f"Merged duplicate (PN captured): {subject_clean} → {case_key}")
            else:
                updates_log.append(f"Merged duplicate: {subject_clean} → {case_key}")
        else:
            upsert_row(row)
            new_threads += 1
            updates_log.append(f"Added new case: {subject_clean} (PN: {pn_final})")
    
    summary = {
        "new": new_threads,
        "updated": updated_threads,
        "filtered_out": filtered_out,
        "unchanged": unchanged,
        "checked": checked,
        "excel_written": (new_threads > 0 or updated_threads > 0),
        "updates_log": updates_log,
    }
    
    if summary["excel_written"]:
        export_to_excel()
    
    update_start_date_env()
    
    return summary

def update_start_date_env():
    """Update START_DATE in .env or session state"""
    now_iso = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    
    # Try to update .env file if it exists (local development)
    env_path = os.path.join(BASE_DIR, ".env")
    if os.path.exists(env_path):
        lines = []
        with open(env_path, "r") as f:
            lines = f.readlines()
        found = False
        new_lines = []
        for line in lines:
            if line.startswith("START_DATE="):
                new_lines.append(f"START_DATE={now_iso}\n")
                found = True
            else:
                new_lines.append(line)
        if not found:
            new_lines.append(f"START_DATE={now_iso}\n")
        with open(env_path, "w") as f:
            f.writelines(new_lines)
        print(f"[INFO] START_DATE updated to {now_iso}")
    else:
        # For cloud deployment, just log it
        print(f"[INFO] Would update START_DATE to {now_iso} (file not writable)")

# REMOVED: All Tkinter functions (update_complaint_log, show_success_popup)
# These are replaced by Streamlit UI in streamlit_app.py

if __name__ == "__main__":
    t0 = time.time()
    try:
        process()
    except Exception as e:
        print(f"[FATAL] {e}")
        sys.exit(1)
    finally:
        dt = time.time() - t0
        print(f"[DONE] Elapsed: {dt:.1f}s")

