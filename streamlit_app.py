"""
MAC Quality Dashboard - Streamlit Interface
Clean design with popup authentication
"""

# ===== PATH FIX FOR EXE DEPLOYMENT =====
import sys
import os
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)
# ===== END PATH FIX =====

# Standard library imports
import sys
import time
import sqlite3
import threading
from datetime import datetime
from typing import Dict, List, Optional
from io import StringIO
from contextlib import redirect_stdout

# Third-party imports
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

# Import from existing modules (all main imports consolidated)
from main import (
    BASE_DIR, process, fetch_all_rows, EXCEL_PATH, DB_PATH,
    to_et_naive, init_db, export_to_excel, app, SCOPES
)
from prompts import CATEGORIES

# ==========================================
# Page Configuration
# ==========================================
st.set_page_config(
    page_title="MAC Quality Dashboard",
    page_icon=os.path.join(BASE_DIR, "mac_logo.png"),
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==========================================
# MAC Blue & White Theme
# ==========================================
PRIMARY_COLOR = "#1E3A8A"
SECONDARY_COLOR = "#EEF2FF"
BACKGROUND_COLOR = "#FFFFFF"
TEXT_COLOR = "#0F172A"
BUTTON_COLOR = "#1E40AF"
CARD_BG = "#FFFFFF"
HEADER_BG = "#1E3A8A"

st.markdown(f"""
<style>
    /* MAC Blue & White Theme */
    :root {{
        --primary-color: {PRIMARY_COLOR};
        --secondary-color: {SECONDARY_COLOR};
        --background-color: {BACKGROUND_COLOR};
        --text-color: {TEXT_COLOR};
    }}
    
    .main {{
        background-color: {BACKGROUND_COLOR};
    }}
    
    .main-header {{
        background-color: {HEADER_BG};
        padding: 2rem;
        border-radius: 0px;
        color: white;
        margin-bottom: 2rem;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
    }}
    
    .main-header h1 {{
        margin: 0;
        font-size: 2.5rem;
        font-weight: 700;
        color: white;
    }}
    
    .main-header p {{
        margin: 0.5rem 0 0 0;
        font-size: 1.1rem;
        opacity: 0.9;
        color: white;
    }}
    
    .stMetric {{
        background: white;
        padding: 1rem;
        border-radius: 8px;
        border: 1px solid {SECONDARY_COLOR};
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
    }}
    
    [data-testid="stSidebar"] {{
        background-color: {SECONDARY_COLOR};
    }}
    
    .stButton>button {{
        background-color: {BUTTON_COLOR};
        color: white;
        border-radius: 6px;
        font-weight: 600;
        border: none;
        transition: all 0.2s ease;
    }}

    .stButton>button:hover {{
        background-color: #1E3A8A;
        transform: translateY(-1px);
        box-shadow: 0 4px 8px rgba(30, 58, 138, 0.3);
    }}

    .stDownloadButton>button {{
        background-color: {PRIMARY_COLOR} !important;
        color: white !important;
        border-radius: 6px;
        font-weight: 600;
        border: none;
        transition: all 0.2s ease;
    }}

    .stDownloadButton>button:hover {{
        background-color: #1E3A8A !important;
        transform: translateY(-1px);
        box-shadow: 0 4px 8px rgba(30, 58, 138, 0.3);
    }}
    
    /* Link Buttons - MAC Blue (for Open Login Page) */
    a[kind="primary"], a[kind="secondary"] {{
        background-color: {BUTTON_COLOR} !important;
        color: white !important;
        border-radius: 6px;
        font-weight: 600;
        border: none;
        text-decoration: none;
        padding: 0.5rem 1rem;
        display: inline-block;
        transition: all 0.2s ease;
    }}
    
    a[kind="primary"]:hover, a[kind="secondary"]:hover {{
        background-color: #1E3A8A !important;
        transform: translateY(-1px);
        box-shadow: 0 4px 8px rgba(30, 58, 138, 0.3);
    }}
    
    /* Multiselect tags - MAC Blue instead of red */
    .stMultiSelect [data-baseweb="tag"] {{
        background-color: {PRIMARY_COLOR} !important;
    }}
    
    .stMultiSelect [data-baseweb="tag"] span {{
        color: white !important;
    }}
    
    .stTabs [data-baseweb="tab-list"] {{
        gap: 8px;
    }}
    
    .stTabs [data-baseweb="tab"] {{
        background-color: white;
        border-radius: 4px;
        padding: 10px 20px;
        color: {TEXT_COLOR};
        border: 1px solid {SECONDARY_COLOR};
    }}
    
    .stTabs [aria-selected="true"] {{
        background-color: {PRIMARY_COLOR};
        color: white;
        border-color: {PRIMARY_COLOR};
    }}
    
    .stSuccess {{
        background-color: #D1FAE5;
        color: #065F46;
        border-left: 4px solid #059669;
    }}
    
    .stInfo {{
        background-color: {SECONDARY_COLOR};
        color: {PRIMARY_COLOR};
        border-left: 4px solid {PRIMARY_COLOR};
    }}
    
    .stWarning {{
        background-color: {SECONDARY_COLOR};
        color: {PRIMARY_COLOR};
        border-left: 4px solid {PRIMARY_COLOR};
    }}
    
    .stError {{
        background-color: {SECONDARY_COLOR};
        color: {PRIMARY_COLOR};
        border-left: 4px solid {PRIMARY_COLOR};
    }}
</style>
""", unsafe_allow_html=True)

# ==========================================
# Session State Initialization
# ==========================================
if 'df' not in st.session_state:
    st.session_state.df = pd.DataFrame()
if 'last_sync' not in st.session_state:
    st.session_state.last_sync = None
if 'sync_running' not in st.session_state:
    st.session_state.sync_running = False
if 'sync_logs' not in st.session_state:
    st.session_state.sync_logs = []
if 'show_auth_dialog' not in st.session_state:
    st.session_state.show_auth_dialog = False
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'show_sync_report' not in st.session_state:
    st.session_state.show_sync_report = False
if 'sync_report_data' not in st.session_state:
    st.session_state.sync_report_data = None
if 'db_downloaded' not in st.session_state:
    st.session_state.db_downloaded = False

# ==========================================
# Helper Functions
# ==========================================

def download_db_from_github() -> bool:
    """Download complaints.db from the data branch on GitHub.
    Requires GITHUB_REPO (e.g. 'owner/repo') in secrets or env.
    For private repos, also set GITHUB_TOKEN.
    """
    try:
        github_repo = ""
        github_token = ""
        try:
            github_repo = st.secrets["credentials"].get("GITHUB_REPO", "")
            github_token = st.secrets["credentials"].get("GITHUB_TOKEN", "")
        except (KeyError, FileNotFoundError, AttributeError):
            pass
        if not github_repo:
            github_repo = os.getenv("GITHUB_REPO", "")
        if not github_token:
            github_token = os.getenv("GITHUB_TOKEN", "")

        if not github_repo:
            return False

        url = f"https://raw.githubusercontent.com/{github_repo}/data/complaints.db"
        headers = {}
        if github_token:
            headers["Authorization"] = f"token {github_token}"

        import requests as req
        resp = req.get(url, headers=headers, timeout=60)
        if resp.status_code == 200:
            with open(DB_PATH, "wb") as f:
                f.write(resp.content)
            print(f"[OK] Downloaded database from GitHub ({len(resp.content):,} bytes)")
            return True
        else:
            print(f"[WARN] Could not download database from GitHub: HTTP {resp.status_code}")
            return False
    except Exception as e:
        print(f"[WARN] Failed to download database from GitHub: {e}")
        return False


def load_custom_columns() -> List[str]:
    """Load custom column definitions from database"""
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS custom_columns (
                column_name TEXT PRIMARY KEY,
                column_type TEXT DEFAULT 'TEXT'
            )
        """)
        cur.execute("SELECT column_name FROM custom_columns")
        cols = [row[0] for row in cur.fetchall()]
        con.close()
        return cols
    except Exception:
        return []

def save_custom_column(col_name: str) -> bool:
    """Save new custom column to database"""
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        cur.execute("INSERT OR IGNORE INTO custom_columns (column_name) VALUES (?)", (col_name,))
        
        cur.execute("PRAGMA table_info(complaints)")
        existing = [row[1] for row in cur.fetchall()]
        if col_name not in existing:
            cur.execute(f"ALTER TABLE complaints ADD COLUMN [{col_name}] TEXT")
        
        con.commit()
        con.close()
        return True
    except Exception as e:
        st.error(f"Failed to add column: {e}")
        return False

def delete_custom_column(col_name: str) -> bool:
    """Remove custom column from tracking"""
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        cur.execute("DELETE FROM custom_columns WHERE column_name=?", (col_name,))
        con.commit()
        con.close()
        return True
    except Exception as e:
        st.error(f"Failed to delete column: {e}")
        return False

def get_setting(key: str, default: str = "") -> str:
    """Get a setting value from the database"""
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        cur.execute("SELECT value FROM settings WHERE key=?", (key,))
        row = cur.fetchone()
        con.close()
        return row[0] if row else default
    except Exception:
        return default

def set_setting(key: str, value: str) -> bool:
    """Save a setting value to the database"""
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        cur.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
        con.commit()
        con.close()
        return True
    except Exception:
        return False

def update_cell_in_db(conversation_id: str, col_name: str, new_value: str):
    """Update a single cell in the database"""
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        
        col_map = {
            "Date (ET)": "first_seen_utc",
            "Initiated By": "initiator_email",
            "P/N": "part_number",
            "Category": "category",
            "Summary": "summary",
            "Subject": "subject",
            "Link": "thread_url"
        }
        
        db_col = col_map.get(col_name, col_name)
        
        cur.execute(f"UPDATE complaints SET [{db_col}]=? WHERE conversation_id=?", 
                   (new_value, conversation_id))
        con.commit()
        con.close()
        return True
    except Exception as e:
        st.error(f"Database error: {e}")
        return False

def delete_row_from_db(conversation_id: str) -> bool:
    """Delete a complaint row from database"""
    try:
        con = sqlite3.connect(DB_PATH)
        cur = con.cursor()
        cur.execute("DELETE FROM complaints WHERE conversation_id=?", (conversation_id,))
        con.commit()
        con.close()
        return True
    except Exception as e:
        st.error(f"Failed to delete: {e}")
        return False

def load_data() -> pd.DataFrame:
    """Load and prepare data from database.
    On first call per session, downloads latest db from GitHub data branch.
    """
    if not st.session_state.get('db_downloaded', False):
        if download_db_from_github():
            st.session_state.db_downloaded = True

    init_db()
    df = fetch_all_rows()

    if df.empty:
        return df

    def to_et_wrapper(x):
        return to_et_naive(x) if pd.notna(x) and str(x).strip() else None

    df["__first_et"] = df["first_seen_utc"].apply(to_et_wrapper) if "first_seen_utc" in df.columns else None

    display = pd.DataFrame()
    display["Date (ET)"] = df["__first_et"]
    display["Initiated By"] = df.get("initiator_email", "")
    display["P/N"] = df.get("part_number", "")
    display["Category"] = df.get("category", "")
    display["Summary"] = df.get("summary", "")
    display["Subject"] = df.get("subject", "")
    display["Link"] = df.get("thread_url", "")

    custom_cols = load_custom_columns()
    for col in custom_cols:
        display[col] = df.get(col, "")

    display["_conversation_id"] = df.get("conversation_id", "")

    return display

def generate_excel_bytes() -> bytes:
    """Generate Excel file as bytes from current database data"""
    from io import BytesIO
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import Alignment, Font

    df = fetch_all_rows()

    if df.empty:
        buffer = BytesIO()
        pd.DataFrame({"Message": ["No data available"]}).to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)
        return buffer.getvalue()

    if "case_key" in df.columns and "received_utc" in df.columns:
        df = df.sort_values("received_utc").drop_duplicates(subset=["case_key"], keep="last")

    def _to_et_naive_wrapper(x):
        return to_et_naive(x) if pd.notna(x) and str(x).strip() else None

    df["__first_et"] = df["first_seen_utc"].apply(_to_et_naive_wrapper)
    df["Date (ET)"] = df["__first_et"]
    df = df.sort_values("__first_et", ascending=False, na_position="last")

    colmap = {
        "initiator_email": "Initiated By",
        "part_number": "P/N",
        "summary": "Summary",
        "category": "Category",
        "subject": "Subject",
        "thread_url": "Link",
    }
    base_cols = ["Date (ET)"] + list(colmap.keys())
    base_cols = [c for c in base_cols if c in df.columns]
    df_out = df[base_cols].rename(columns=colmap)

    if "Category_Final" not in df_out.columns:
        if "Category" in df_out.columns:
            df_out.insert(df_out.columns.get_loc("Category") + 1, "Category_Final", "")
        else:
            df_out["Category_Final"] = ""

    if "Notes" not in df_out.columns:
        if "Subject" in df_out.columns:
            df_out.insert(df_out.columns.get_loc("Subject") + 1, "Notes", "")
        else:
            df_out["Notes"] = ""

    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("SELECT column_name FROM custom_columns")
    custom_cols = [row[0] for row in cur.fetchall()]
    con.close()

    for col in custom_cols:
        if col in df.columns and col not in df_out.columns:
            df_out[col] = df[col]

    if "Link" in df_out.columns:
        df_out["_url"] = df_out["Link"].copy()
        df_out["Link"] = df_out["_url"].apply(
            lambda u: f'=HYPERLINK("{str(u).strip()}", "Open")' if pd.notna(u) and str(u).strip() else ""
        )

    desired = [
        "Date (ET)",
        "Initiated By",
        "P/N",
        "Category", "Category_Final",
        "Summary",
        "Subject",
        "Notes",
        "Link",
    ] + custom_cols
    existing = [c for c in desired if c in df_out.columns]
    if "_url" in df_out.columns:
        df_out = df_out.drop(columns=["_url"])
    df_out = df_out[existing]

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet = "Complaints"
        df_out.to_excel(writer, sheet_name=sheet, index=False)
        wb = writer.book
        ws = wb[sheet]
        ws.freeze_panes = "A2"
        last_col = get_column_letter(ws.max_column)
        last_row = ws.max_row
        tbl = Table(displayName="ComplaintTable", ref=f"A1:{last_col}{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tbl)
        if "Summary" in df_out.columns:
            cidx = df_out.columns.get_loc("Summary") + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=cidx).alignment = Alignment(wrap_text=True, vertical="top")
        if "Date (ET)" in df_out.columns:
            didx = df_out.columns.get_loc("Date (ET)") + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=didx).number_format = "mm/dd/yyyy"
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        target_widths = {
            "Date (ET)": 12,
            "Initiated By": 30,
            "P/N": 26,
            "Summary": 64,
            "Category": 18, "Category_Final": 18,
            "Subject": 44,
            "Notes": 30,
            "Link": 10,
        }
        for idx, name in enumerate(df_out.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = target_widths.get(name, 24)

    buffer.seek(0)
    return buffer.getvalue()

def log_message(msg: str):
    """Add message to sync logs with timestamp"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    log_entry = f"[{timestamp}] {msg}\n"
    st.session_state.sync_logs.append(log_entry)
    print(log_entry.strip())

def run_sync_process(ui_log_callback=None):
    """Run the email sync process (main.py process() function)
    ui_log_callback: optional function(msg) to show real-time logs in UI
    """
    def combined_log(msg):
        """Log to both session state and UI callback"""
        log_message(msg)
        if ui_log_callback:
            try:
                ui_log_callback(msg)
            except:
                pass

    try:
        st.session_state.sync_running = True
        combined_log("="*60)
        combined_log("Starting Email Sync Process")
        combined_log("="*60)

        # Use START_DATE from secrets/env directly (like the original dashboard)
        # This ensures we always sync from the configured start date
        from main import START_DATE
        combined_log(f"Syncing emails since: {START_DATE}")

        combined_log("Calling process() function from main.py...")
        combined_log("This may take 1-3 minutes depending on email volume...")

        # Call the main process function WITHOUT override - uses START_DATE from main.py
        # This matches the original dashboard.py behavior
        # IMPORTANT: Pass combined_log (not ui_log_callback) so logs are saved to session state
        summary = process(log_callback=combined_log)

        log_message("process() completed successfully")

        if summary:
            log_message("\nSync Summary:")
            log_message(f"  - Total emails checked: {summary.get('checked', 0)}")
            log_message(f"  - Filtered (no keywords): {summary.get('filtered_noise', 0)}")
            log_message(f"  - Filtered (Gemini said not complaint): {summary.get('filtered_not_complaint', 0)}")
            log_message(f"  - Unchanged (already in DB): {summary.get('unchanged', 0)}")
            log_message(f"  - New complaints: {summary.get('new', 0)}")
            log_message(f"  - Updated: {summary.get('updated', 0)}")
        else:
            log_message("WARNING: process() returned None or empty summary")
            summary = {"new": 0, "updated": 0, "filtered_out": 0, "unchanged": 0, "checked": 0, "updates_log": []}

        # Note: We don't update last_sync_date - each sync uses START_DATE from secrets
        # This matches the original dashboard.py behavior

        log_message("="*60)
        log_message("Sync completed successfully!")
        log_message("="*60)

        st.session_state.last_sync = datetime.now()
        st.session_state.sync_summary = summary

        log_message("Loading updated data from database...")
        st.session_state.df = load_data()
        log_message("Data loaded successfully")

        st.session_state.sync_running = False
        return summary
    except Exception as e:
        import traceback
        error_msg = f"FATAL ERROR: {str(e)}\nTraceback:\n{traceback.format_exc()}"
        log_message(error_msg)
        st.session_state.sync_running = False
        raise e

def check_authentication() -> bool:
    """Check if user is authenticated"""
    if "access_token" in st.session_state and "token_expires_at" in st.session_state:
        if time.time() < st.session_state.token_expires_at:
            return True
    
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
        if result and "access_token" in result:
            st.session_state.access_token = result["access_token"]
            st.session_state.token_expires_at = time.time() + result.get("expires_in", 3600)
            st.session_state.authenticated = True
            return True
    
    return False

@st.dialog("Sync Report", width="medium")
def show_sync_report_dialog():
    """Display sync report in a modal dialog"""
    summary = st.session_state.sync_report_data
    
    st.success("Sync completed successfully!")
    
    # Show metrics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("New Complaints", summary.get('new', 0))
    with col2:
        st.metric("Updated", summary.get('updated', 0))
    with col3:
        st.metric("Filtered Out", summary.get('filtered_out', 0))
    with col4:
        st.metric("Total Checked", summary.get('checked', 0))
    
    # Show recent changes
    if summary.get('updates_log'):
        st.markdown("---")
        st.markdown("**Recent Changes:**")
        for log_entry in summary['updates_log'][:10]:  # Show last 10
            st.caption(log_entry)
    
    st.markdown("---")
    
    # Close button
    if st.button("Close", use_container_width=True, type="primary"):
        st.session_state.show_sync_report = False
        st.session_state.sync_report_data = None
        st.session_state.df = load_data()  # Reload data
        st.rerun()

@st.dialog("Microsoft Authentication Required", width="medium")
def show_auth_dialog():
    """Display authentication dialog"""
    st.warning("To sync emails, sign in with your Microsoft account")
    
    # Start device flow if not already started
    if "device_flow" not in st.session_state:
        try:
            flow = app.initiate_device_flow(scopes=SCOPES)
            if "user_code" in flow:
                st.session_state.device_flow = flow
                st.session_state.auth_started = time.time()
            else:
                st.error(f"Failed to start authentication: {flow.get('error_description', 'Unknown error')}")
                return
        except Exception as e:
            st.error(f"Authentication error: {str(e)}")
            return
    
    flow = st.session_state.device_flow
    
    st.markdown("**Your authentication code:**")
    st.code(flow["user_code"], language=None)
    
    col1, col2 = st.columns(2)
    with col1:
        st.link_button("Open Login Page", "https://microsoft.com/devicelogin", use_container_width=True)
    with col2:
        if st.button("Run Sync", use_container_width=True, type="primary"):
            try:
                log_message("Run Sync button clicked from auth dialog")
                result = app.acquire_token_by_device_flow(flow)
                
                if "access_token" in result:
                    st.session_state.access_token = result["access_token"]
                    st.session_state.token_expires_at = time.time() + result.get("expires_in", 3600)
                    st.session_state.authenticated = True
                    st.session_state.pop("device_flow", None)
                    st.session_state.pop("auth_started", None)
                    st.session_state.show_auth_dialog = False
                    
                    log_message("Authentication successful! Starting sync...")
                    st.success("Authenticated! Starting sync...")
                    
                    # Run sync with progress indicator
                    try:
                        summary = run_sync_process()
                        log_message("Sync completed, showing report")
                        
                        # Show sync report
                        st.session_state.show_sync_report = True
                        st.session_state.sync_report_data = summary
                        st.rerun()
                    except Exception as sync_error:
                        log_message(f"Sync failed: {str(sync_error)}")
                        st.error(f"Sync failed: {str(sync_error)}")
                        # Close dialog even if sync fails
                        time.sleep(3)
                        st.rerun()
                else:
                    if "pending" in result.get('error_description', '').lower():
                        st.warning("Complete sign-in first, then click 'Run Sync'")
                        log_message("Auth still pending")
                    else:
                        error_msg = result.get('error_description', 'Unknown error')
                        st.error(f"Authentication failed: {error_msg}")
                        log_message(f"Auth failed: {error_msg}")
            except Exception as e:
                error_msg = str(e)
                st.error(f"Error: {error_msg}")
                log_message(f"Error in auth/sync: {error_msg}")
                import traceback
                log_message(traceback.format_exc())
    
    # Show timeout
    if "auth_started" in st.session_state:
        elapsed = time.time() - st.session_state.auth_started
        remaining = max(0, 900 - elapsed)
        if remaining > 0:
            minutes = int(remaining // 60)
            seconds = int(remaining % 60)
            st.caption(f"Code expires in {minutes}m {seconds}s")
        else:
            st.error("Code expired!")
            if st.button("Get New Code"):
                st.session_state.pop("device_flow", None)
                st.session_state.pop("auth_started", None)
                st.rerun()

# ==========================================
# Header Section
# ==========================================
st.markdown(f"""
<div class="main-header">
    <h1>MAC PRODUCTS</h1>
    <p>Quality Automation Dashboard - Complaint Management System</p>
</div>
""", unsafe_allow_html=True)

# ==========================================
# Show Dialogs
# ==========================================
if st.session_state.show_sync_report and st.session_state.sync_report_data:
    show_sync_report_dialog()

if st.session_state.show_auth_dialog:
    show_auth_dialog()

# ==========================================
# Sidebar - Clean Controls
# ==========================================
with st.sidebar:
    logo_path = os.path.join(BASE_DIR, "mac_logo.png")
    if os.path.exists(logo_path):
        st.image(logo_path, use_container_width=True)
    
    st.markdown("---")
    st.header("Controls")
    
    # RUN EMAIL SYNC - Main button
    if st.button("Run Email Sync", use_container_width=True, type="primary"):
        if check_authentication():
            # Already authenticated, just run sync
            log_message("Run Email Sync clicked - already authenticated")
            try:
                # Use st.status() for real-time progress updates
                with st.status("Syncing emails...", expanded=True) as status:
                    status.write("Starting sync process...")

                    # Create callback to show logs in real-time
                    def ui_log(msg):
                        status.write(msg)

                    # Run sync with UI callback for real-time logs
                    summary = run_sync_process(ui_log_callback=ui_log)

                    # Show final results
                    if summary:
                        status.write("---")
                        status.write(f"FINAL: Checked {summary.get('checked', 0)} emails")
                        status.write(f"FINAL: New complaints: {summary.get('new', 0)}")
                        status.write(f"FINAL: Updated: {summary.get('updated', 0)}")
                        status.write(f"FINAL: Filtered out: {summary.get('filtered_out', 0)}")
                        status.update(label="Sync completed!", state="complete", expanded=True)
                    else:
                        status.update(label="Sync completed with no results", state="complete", expanded=True)

                # Store summary in session state to display it
                st.session_state.show_sync_report = True
                st.session_state.sync_report_data = summary
                st.rerun()

            except Exception as e:
                st.error(f"Sync failed: {str(e)}")
                log_message(f"ERROR: {str(e)}")
        else:
            # Need authentication, show dialog
            st.session_state.show_auth_dialog = True
            st.rerun()
    
    # REFRESH DATA (re-downloads from GitHub)
    if st.button("Refresh Data", use_container_width=True):
        st.session_state.db_downloaded = False  # Force re-download
        st.session_state.df = load_data()
        st.rerun()
    
    # DOWNLOAD EXCEL
    try:
        excel_data = generate_excel_bytes()
        st.download_button(
            label="Download Excel",
            data=excel_data,
            file_name=f"Complaint_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    except Exception as e:
        st.error(f"Failed to generate Excel: {e}")

    # Editable sync start date
    last_sync_date = get_setting("last_sync_date", "2025-01-01T00:00:00Z")
    try:
        current_date = datetime.strptime(last_sync_date[:10], "%Y-%m-%d").date()
    except:
        current_date = datetime(2025, 1, 1).date()

    new_date = st.date_input(
        "Next sync starts from",
        value=current_date,
        help="Change this to re-sync emails from a different date"
    )

    # Save if date changed
    if new_date != current_date:
        new_date_iso = new_date.strftime("%Y-%m-%dT00:00:00Z")
        set_setting("last_sync_date", new_date_iso)
        st.rerun()

    st.markdown("---")
    st.header("Filters")
    
    if st.session_state.df.empty:
        st.session_state.df = load_data()
    
    df = st.session_state.df
    
    if not df.empty and "Category" in df.columns:
        categories = ["(All)"] + sorted(df["Category"].dropna().unique().tolist())
        category_filter = st.selectbox("Category", categories, key="category_filter")
    else:
        category_filter = "(All)"
    
    pn_filter = st.text_input("Part Number", placeholder="Search P/N...", key="pn_filter")
    initiator_filter = st.text_input("Initiated By", placeholder="Search email...", key="initiator_filter")
    subject_filter = st.text_input("Subject", placeholder="Search subject...", key="subject_filter")
    
    st.markdown("**Date Range**")
    if not df.empty and "Date (ET)" in df.columns:
        dates = pd.to_datetime(df["Date (ET)"], errors='coerce').dropna()
        if not dates.empty:
            min_date = dates.min().date()
            max_date = dates.max().date()
            date_range = st.date_input(
                "Select range",
                value=(min_date, max_date),
                min_value=min_date,
                max_value=max_date,
                key="date_range"
            )
        else:
            date_range = None
    else:
        date_range = None
    
    st.markdown("---")
    st.header("Column Management")
    
    with st.expander("Add Custom Column"):
        new_col_name = st.text_input("Column Name", key="new_col_input")
        if st.button("Add Column", key="add_col_btn"):
            if new_col_name and new_col_name.strip():
                if save_custom_column(new_col_name.strip()):
                    st.success(f"Added '{new_col_name}'")
                    st.session_state.df = load_data()
                    st.rerun()
            else:
                st.warning("Please enter a column name")
    
    custom_cols = load_custom_columns()
    if custom_cols:
        with st.expander("Delete Custom Column"):
            col_to_delete = st.selectbox("Select column", custom_cols, key="del_col_select")
            if st.button("Delete Column", key="del_col_btn"):
                if delete_custom_column(col_to_delete):
                    st.success(f"Deleted '{col_to_delete}'")
                    st.session_state.df = load_data()
                    st.rerun()
    
    st.markdown("---")
    if st.session_state.last_sync:
        st.caption(f"Last sync: {st.session_state.last_sync.strftime('%Y-%m-%d %H:%M:%S')}")

# ==========================================
# Main Content Area
# ==========================================

df = st.session_state.df

if df.empty:
    st.info("No complaints found. Click 'Run Email Sync' to fetch data.")

    # Show System Logs even when no data - so users can see what happened during sync
    st.markdown("---")
    st.subheader("System Logs")
    st.info("Real-time logs from email sync operations.")

    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button("Clear Logs", type="secondary", key="clear_logs_empty"):
            st.session_state.sync_logs = []
            st.rerun()

    if st.session_state.sync_logs:
        all_logs = "".join(st.session_state.sync_logs)
        log_count = len(st.session_state.sync_logs)
        st.caption(f"Total log entries: {log_count}")

        st.text_area("Logs", value=all_logs, height=400, disabled=True, label_visibility="collapsed")

        st.download_button(
            label="Download Logs",
            data=all_logs,
            file_name=f"sync_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
            key="download_logs_empty"
        )
    else:
        st.info("No logs yet. Click 'Run Email Sync' to see what happens during the sync process.")
else:
    # Apply filters
    filtered_df = df.copy()
    
    if category_filter != "(All)":
        filtered_df = filtered_df[filtered_df["Category"] == category_filter]
    
    if pn_filter.strip():
        filtered_df = filtered_df[
            filtered_df["P/N"].astype(str).str.upper().str.contains(pn_filter.strip().upper(), na=False)
        ]
    
    if initiator_filter.strip():
        filtered_df = filtered_df[
            filtered_df["Initiated By"].astype(str).str.lower().str.contains(initiator_filter.strip().lower(), na=False)
        ]
    
    if subject_filter.strip():
        filtered_df = filtered_df[
            filtered_df["Subject"].astype(str).str.lower().str.contains(subject_filter.strip().lower(), na=False)
        ]
    
    if date_range and len(date_range) == 2:
        filtered_df["Date (ET)"] = pd.to_datetime(filtered_df["Date (ET)"], errors='coerce')
        start_date, end_date = date_range
        filtered_df = filtered_df[
            (filtered_df["Date (ET)"].dt.date >= start_date) &
            (filtered_df["Date (ET)"].dt.date <= end_date)
        ]
    
    # Metrics Row
    st.subheader("Dashboard Metrics")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.metric(label="Total Complaints", value=len(df))
    
    with col2:
        st.metric(label="Displayed", value=len(filtered_df))
    
    with col3:
        if "Category" in filtered_df.columns:
            unique_categories = filtered_df["Category"].nunique()
            st.metric(label="Categories", value=unique_categories)
    
    with col4:
        if "P/N" in filtered_df.columns:
            unique_pns = filtered_df[filtered_df["P/N"] != "No part number provided"]["P/N"].nunique()
            st.metric(label="Unique P/Ns", value=unique_pns)
    
    with col5:
        if "Date (ET)" in filtered_df.columns:
            recent = filtered_df[pd.to_datetime(filtered_df["Date (ET)"], errors='coerce') > 
                                (datetime.now() - pd.Timedelta(days=30))]
            st.metric(label="Last 30 Days", value=len(recent))
    
    # Tabs
    tab1, tab2, tab3, tab4 = st.tabs(["Data Table", "Analytics", "Category Breakdown", "System Logs"])
    
    with tab1:
        st.subheader("Complaint Records")
        st.info("Click any cell to edit. Changes are saved automatically.")
        
        display_columns = [col for col in filtered_df.columns if not col.startswith("_")]
        
        selected_columns = st.multiselect(
            "Select columns to display",
            display_columns,
            default=[col for col in display_columns if col not in ["Link"]],
            key="col_selector"
        )
        
        if selected_columns:
            display_df = filtered_df[selected_columns + ["_conversation_id"]].copy()
            
            col1, col2 = st.columns([3, 1])
            with col1:
                sort_by = st.selectbox("Sort by", selected_columns, key="sort_by")
            with col2:
                sort_order = st.selectbox("Order", ["Descending", "Ascending"], key="sort_order")
            
            if sort_by:
                ascending = (sort_order == "Ascending")
                display_df = display_df.sort_values(sort_by, ascending=ascending, na_position="last")
            
            edited_df = st.data_editor(
                display_df.drop(columns=["_conversation_id"]),
                use_container_width=True,
                height=500,
                num_rows="fixed",
                column_config={
                    "Date (ET)": st.column_config.DatetimeColumn("Date (ET)", format="YYYY-MM-DD HH:mm:ss", disabled=True),
                    "Link": st.column_config.LinkColumn("Link", disabled=True),
                    "Summary": st.column_config.TextColumn("Summary", width="large", max_chars=500),
                    "P/N": st.column_config.TextColumn("P/N", max_chars=50),
                    "Category": st.column_config.SelectboxColumn("Category", options=CATEGORIES, required=True),
                    "Subject": st.column_config.TextColumn("Subject", width="medium"),
                    "Initiated By": st.column_config.TextColumn("Initiated By")
                },
                key="data_editor"
            )
            
            if not edited_df.equals(display_df.drop(columns=["_conversation_id"])):
                st.warning("Detected changes. Click below to save.")
                col1, col2 = st.columns([1, 4])
                with col1:
                    if st.button("Save Changes", type="primary"):
                        changes_saved = 0
                        for idx in edited_df.index:
                            if idx in display_df.index:
                                conv_id = display_df.loc[idx, "_conversation_id"]
                                for col in selected_columns:
                                    if col in edited_df.columns:
                                        old_val = str(display_df.loc[idx, col])
                                        new_val = str(edited_df.loc[idx, col])
                                        if old_val != new_val:
                                            if update_cell_in_db(conv_id, col, new_val):
                                                changes_saved += 1
                        
                        if changes_saved > 0:
                            st.success(f"Saved {changes_saved} changes!")
                            st.session_state.df = load_data()
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.info("No changes to save")
                with col2:
                    if st.button("Cancel", type="secondary"):
                        st.rerun()
            
            st.markdown("---")
            st.markdown("**Delete a Record**")
            row_to_delete = st.selectbox(
                "Select row to delete (by Subject)",
                range(len(display_df)),
                format_func=lambda i: f"{i}: {display_df.iloc[i]['Subject'][:60]}...",
                key="delete_selector"
            )
            
            if st.button("Delete Selected Row", type="secondary"):
                conv_id = display_df.iloc[row_to_delete]["_conversation_id"]
                if delete_row_from_db(conv_id):
                    st.success("Record deleted!")
                    st.session_state.df = load_data()
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("Failed to delete")
            
            st.markdown("---")
            csv = display_df.drop(columns=["_conversation_id"]).to_csv(index=False)
            st.download_button(
                label="Download Filtered Data (CSV)",
                data=csv,
                file_name=f"complaints_filtered_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )
    
    with tab2:
        st.subheader("Complaint Trends")
        
        if "Date (ET)" in filtered_df.columns and not filtered_df.empty:
            df_with_date = filtered_df.copy()
            df_with_date["Date (ET)"] = pd.to_datetime(df_with_date["Date (ET)"], errors='coerce')
            df_with_date = df_with_date.dropna(subset=["Date (ET)"])
            
            if not df_with_date.empty:
                df_with_date["Month"] = df_with_date["Date (ET)"].dt.to_period("M").astype(str)
                monthly_counts = df_with_date.groupby("Month").size().reset_index(name="Count")
                
                fig = px.line(monthly_counts, x="Month", y="Count", title="Complaints Over Time", markers=True, color_discrete_sequence=[PRIMARY_COLOR])
                fig.update_layout(xaxis_title="Month", yaxis_title="Number of Complaints", hovermode="x unified", plot_bgcolor='white', paper_bgcolor='white', font=dict(color=TEXT_COLOR))
                st.plotly_chart(fig, use_container_width=True)
                
                df_with_date["Week"] = df_with_date["Date (ET)"].dt.to_period("W").astype(str)
                weekly_counts = df_with_date.groupby("Week").size().reset_index(name="Count")
                
                fig2 = px.bar(weekly_counts.tail(12), x="Week", y="Count", title="Last 12 Weeks Activity", color_discrete_sequence=[PRIMARY_COLOR])
                fig2.update_layout(xaxis_title="Week", yaxis_title="Complaints", plot_bgcolor='white', paper_bgcolor='white', font=dict(color=TEXT_COLOR))
                st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("No date information available for trend analysis")
    
    with tab3:
        st.subheader("Category Analysis")
        
        if "Category" in filtered_df.columns and not filtered_df.empty:
            col1, col2 = st.columns(2)
            
            with col1:
                category_counts = filtered_df["Category"].value_counts()
                fig = px.pie(values=category_counts.values, names=category_counts.index, title="Complaint Distribution by Category", color_discrete_sequence=px.colors.sequential.Blues_r)
                fig.update_layout(plot_bgcolor='white', paper_bgcolor='white', font=dict(color=TEXT_COLOR))
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                fig = px.bar(x=category_counts.index, y=category_counts.values, labels={"x": "Category", "y": "Count"}, title="Category Breakdown", color_discrete_sequence=[PRIMARY_COLOR])
                fig.update_layout(showlegend=False, plot_bgcolor='white', paper_bgcolor='white', font=dict(color=TEXT_COLOR))
                st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("**Category Statistics**")
            category_stats = pd.DataFrame({
                "Category": category_counts.index,
                "Count": category_counts.values,
                "Percentage": (category_counts.values / category_counts.sum() * 100).round(2)
            })
            st.dataframe(category_stats, use_container_width=True, hide_index=True)
        else:
            st.info("No category information available")
    
    with tab4:
        st.subheader("System Logs")
        st.info("Real-time logs from email sync operations.")
        
        col1, col2 = st.columns([1, 4])
        with col1:
            if st.button("Clear Logs", type="secondary"):
                st.session_state.sync_logs = []
                st.rerun()
        
        if st.session_state.sync_logs:
            all_logs = "".join(st.session_state.sync_logs)
            log_count = len(st.session_state.sync_logs)
            st.caption(f"Total log entries: {log_count}")
            
            st.text_area("Logs", value=all_logs, height=600, disabled=True, label_visibility="collapsed")
            
            st.download_button(
                label="Download Logs",
                data=all_logs,
                file_name=f"sync_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime="text/plain"
            )
        else:
            st.info("No logs yet. Run 'Run Email Sync' to generate logs.")

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #64748B; padding: 1rem;">
    <p>MAC Quality Dashboard v2.1 | Powered by Streamlit | © 2025</p>
</div>
""", unsafe_allow_html=True)
